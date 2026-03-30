"""Excel spreadsheet parser using openpyxl."""

import io
import logging
from typing import Any, Optional

from agent_rag.ingestion.parsing.base import BaseParser, ParsedDocument
from agent_rag.ingestion.parsing.utils import normalize_text

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    """
    Parser for Excel spreadsheet files.

    Uses openpyxl for .xlsx files.

    Supports:
    - .xlsx, .xlsm files
    - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet

    Priority: 10 (default for office documents)
    """

    def supports(self, source_type: str, extension: str, mime_type: str) -> bool:
        """Support .xlsx and .xlsm files."""
        if extension.lower() in ["xlsx", "xlsm", "xls"]:
            return True
        if mime_type and "spreadsheetml" in mime_type:
            return True
        return False

    def parse(self, content: bytes, filename: str) -> ParsedDocument:
        """
        Parse Excel file.

        Args:
            content: Raw Excel bytes
            filename: Original filename

        Returns:
            ParsedDocument with sheet content as text and tables

        Raises:
            ImportError: If openpyxl not installed
            Exception: If parsing fails
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "openpyxl not installed. Run: pip install openpyxl"
            )

        # Load workbook
        file_stream = io.BytesIO(content)

        try:
            workbook = openpyxl.load_workbook(file_stream, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise RuntimeError(f"Failed to parse Excel: {e}")

        # Extract content from all sheets
        text_parts = []
        tables = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract sheet as table
            table_data = self._extract_sheet_as_table(sheet, sheet_name)
            if table_data["rows"]:
                tables.append(table_data)

            # Convert table to markdown text
            sheet_text = self._table_to_text(table_data)
            if sheet_text:
                text_parts.append(f"## {sheet_name}\n\n{sheet_text}")

        # Combine all sheets
        text = "\n\n".join(text_parts)
        text = normalize_text(text)

        # Extract metadata
        metadata = {
            "filename": filename,
            "source_type": "file",
            "format": "xlsx",
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
        }

        # Try to extract workbook properties
        if hasattr(workbook, "properties"):
            props = workbook.properties

            if props.title:
                metadata["title"] = props.title
            if props.creator:
                metadata["author"] = props.creator
            if props.created:
                try:
                    metadata["created"] = props.created.isoformat()
                except Exception:
                    pass
            if props.modified:
                try:
                    metadata["modified"] = props.modified.isoformat()
                except Exception:
                    pass

        return ParsedDocument(
            text=text,
            metadata=metadata,
            images=[],
            links=[],
            tables=tables,
        )

    def _extract_sheet_as_table(
        self,
        sheet: Any,
        sheet_name: str,
    ) -> dict[str, Any]:
        """
        Extract sheet as table structure.

        Args:
            sheet: openpyxl worksheet
            sheet_name: Sheet name

        Returns:
            Table dictionary with rows and metadata
        """
        rows = []

        # Iterate through rows
        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            # Convert row to strings
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell).strip())

            rows.append(row_data)

        # Calculate dimensions
        row_count = len(rows)
        column_count = max((len(row) for row in rows), default=0)

        return {
            "sheet_name": sheet_name,
            "rows": rows,
            "row_count": row_count,
            "column_count": column_count,
        }

    def _table_to_text(self, table_data: dict[str, Any]) -> str:
        """
        Convert table to markdown-style text.

        Args:
            table_data: Table dictionary

        Returns:
            Markdown table text
        """
        rows = table_data["rows"]

        if not rows:
            return ""

        text_parts = []

        # Assume first row is header
        if len(rows) > 0:
            header = rows[0]
            text_parts.append("| " + " | ".join(header) + " |")

            # Add separator
            text_parts.append("| " + " | ".join(["---"] * len(header)) + " |")

            # Add data rows
            for row in rows[1:]:
                # Pad row to match header length
                padded_row = row + [""] * (len(header) - len(row))
                text_parts.append("| " + " | ".join(padded_row[:len(header)]) + " |")

        return "\n".join(text_parts)

    @property
    def priority(self) -> int:
        """Default priority for office documents."""
        return 10


class CSVParser(BaseParser):
    """
    Parser for CSV files.

    Simple CSV parsing using Python's built-in csv module.

    Supports:
    - .csv files
    - text/csv

    Priority: 5 (lower than Excel but higher than plain text)
    """

    def supports(self, source_type: str, extension: str, mime_type: str) -> bool:
        """Support .csv files."""
        if extension.lower() == "csv":
            return True
        if mime_type and "csv" in mime_type.lower():
            return True
        return False

    def parse(self, content: bytes, filename: str) -> ParsedDocument:
        """
        Parse CSV file.

        Args:
            content: Raw CSV bytes
            filename: Original filename

        Returns:
            ParsedDocument with CSV as table and text
        """
        import csv

        # Decode content
        try:
            text_content = content.decode("utf-8")
        except UnicodeDecodeError:
            text_content = content.decode("latin-1")

        # Parse CSV
        reader = csv.reader(text_content.splitlines())
        rows = list(reader)

        # Create table structure
        table_data = {
            "sheet_name": "CSV",
            "rows": rows,
            "row_count": len(rows),
            "column_count": max((len(row) for row in rows), default=0),
        }

        # Convert to text (markdown table)
        text = self._table_to_text(table_data)

        # Metadata
        metadata = {
            "filename": filename,
            "source_type": "file",
            "format": "csv",
            "row_count": len(rows),
        }

        return ParsedDocument(
            text=text,
            metadata=metadata,
            images=[],
            links=[],
            tables=[table_data],
        )

    def _table_to_text(self, table_data: dict[str, Any]) -> str:
        """Convert CSV table to markdown."""
        rows = table_data["rows"]

        if not rows:
            return ""

        text_parts = []

        # Header row
        if len(rows) > 0:
            header = rows[0]
            text_parts.append("| " + " | ".join(header) + " |")
            text_parts.append("| " + " | ".join(["---"] * len(header)) + " |")

            # Data rows
            for row in rows[1:]:
                padded_row = row + [""] * (len(header) - len(row))
                text_parts.append("| " + " | ".join(padded_row[:len(header)]) + " |")

        return "\n".join(text_parts)

    @property
    def priority(self) -> int:
        """Lower than Excel but higher than plain text."""
        return 5
